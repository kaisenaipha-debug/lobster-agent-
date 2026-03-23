#!/usr/bin/env python3
"""发货清单生成器 - 像素级还原博敏电子模板"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date

FONT = Font(name='宋体', size=12)
TITLE_FONT = Font(name='宋体', size=18)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


def generate_delivery_list(output_path, client_name, order_no, items,
                           our_company=None, ship_date=None):
    if our_company is None:
        our_company = "醴陵韶峰服饰销售有限公司"
    if ship_date is None:
        ship_date = date.today().strftime("%Y年%m月%d日")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet2"

    # 列宽 - 还原原始模板
    ws.column_dimensions['A'].width = 7.04
    ws.column_dimensions['B'].width = 16.84
    ws.column_dimensions['C'].width = 15.19
    ws.column_dimensions['D'].width = 31.62
    ws.column_dimensions['E'].width = 13.0
    ws.column_dimensions['F'].width = 13.0
    ws.column_dimensions['G'].width = 22.41

    # 行1: 标题 "发货清单"
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = "发货清单"
    c.font = TITLE_FONT
    c.alignment = CENTER
    ws.row_dimensions[1].height = 24.0

    # 行2: 定制方 + 订单编号
    ws.merge_cells('A2:G2')
    c = ws['A2']
    c.value = f"定制方：{client_name}                                                                          订单编号:  {order_no}"
    c.font = FONT
    c.alignment = LEFT
    ws.row_dimensions[2].height = 21.0

    # 行3: 列标题
    headers = ['序号', '物料编码', '物料名称', '物料描述', '单位', '数量', '备注']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = FONT
        c.alignment = CENTER
    ws.row_dimensions[3].height = 24.0

    # 数据行
    row = 4
    for item in items:
        idx, code, name, desc, unit, qty = item[0], item[1], item[2], item[3], item[4], item[5]
        vals = [idx, code, name, desc, unit, qty]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = FONT
            c.alignment = CENTER
        ws.row_dimensions[row].height = 21.0
        row += 1

    # 行5: 承揽方 - 右对齐
    ws.merge_cells(f'A{row}:G{row}')
    c = ws[f'A{row}']
    c.value = f"承揽方：{our_company}"
    c.font = FONT
    c.alignment = RIGHT
    ws.row_dimensions[row].height = 26.0
    row += 1

    # 行6: 发货日期 - 右对齐
    ws.merge_cells(f'A{row}:G{row}')
    c = ws[f'A{row}']
    c.value = f"发货日期：{ship_date}"
    c.font = FONT
    c.alignment = RIGHT
    ws.row_dimensions[row].height = 21.0

    wb.save(output_path)
    print(f"✅ 发货清单已生成: {output_path}")


if __name__ == "__main__":
    test_items = [
        (1, 'N10770333', '短袖衬衫', '男(女）装 竹纤维、抗皱、白色', '件', 6),
    ]
    generate_delivery_list(
        '/tmp/test_delivery.xlsx',
        '博敏电子股份有限公司（一厂）',
        'MZPOM12504290038',
        test_items,
        ship_date='2025年5月20日'
    )
